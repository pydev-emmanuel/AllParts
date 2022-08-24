import os
import xlsxwriter
from openpyxl import load_workbook
from bs4 import BeautifulSoup as soup


def product_data(html_file, code):
    product_details = {}
    oem_equivalent = []
    html = open(html_file, "r")
    contents = html.read()
    bs_content = soup(contents, "lxml")
    price = bs_content.find(class_="quantity quantity--pricesmall productpricetoggle__gross productpricetoggle__wholesale js-product-wholesale-toggle").findNext("div").text
    price = price.replace(".", "")
    price = price.replace(",", ".")
    price = int(float(price))
    price_intercar = price + 30
    price_intercar = int(price_intercar)
    product_details["price"] = price_intercar
    try:
        for tag in bs_content.find_all("li", class_="refnumbers__item"):
            if tag.find('span', class_='refnumbers__manufacturer').text == "Echivalente Inter Cars":
                pass
            else:
                manufacturer = tag.find('span', class_='refnumbers__manufacturer').text
                refnumber = tag.find('span', class_='refnumbers__refnumber').text
                refnumber_mod = refnumber.replace(" ", "")
                oem_equivalent.append(f"{manufacturer} - {refnumber_mod}")
    except UnboundLocalError:
        pass
    except AttributeError:
        pass
    product_details["oem_equivalent"] = oem_equivalent
    try:
        img_src = bs_content.find("img")["src"]
        img_src = img_src.replace("t_t300x300v2/", "")
    except TypeError:
        img_src = None
    product_details["img_src"] = img_src
    descriere_tehnica = bs_content.find("div", class_="producttechnicaldesc producttechnicaldesc--productinfo").text
    descriere_tehnica = descriere_tehnica.replace("\n", "")
    descriere_tehnica = descriere_tehnica.strip()
    descriere_tehnica_list = list(descriere_tehnica)
    try:
        descriere_tehnica_list.remove("È")
        for item in descriere_tehnica_list:
            if item == "™":
                index = descriere_tehnica_list.index(item)
                descriere_tehnica_list[index] = "s"
    except ValueError:
        pass
    descriere_tehnica = "".join(descriere_tehnica_list)
    product_details["descriere_tehnica"] = descriere_tehnica
    tip_produs = bs_content.find("body").p.text
    tip_produs_mod = []
    for x in list(tip_produs):
        if x != "\n":
            tip_produs_mod.append(x)
        else:
            break
    tip_produs = "".join(tip_produs_mod)
    print(tip_produs)
    product_details["tip_produs"] = tip_produs
    workbook = load_workbook("C:\\Users\\HP\\Desktop\\ALLPARTS\\VOLANT\\cod_producator_link.xlsx")
    worksheet = workbook["Sheet1"]
    column_producator = worksheet["B"]
    lista_producator = [column_producator[x].value for x in range(len(column_producator))]
    column_code = worksheet["A"]
    code_list = [column_code[x].value for x in range(len(column_code))]
    for cod_piesa in code_list:
        if code == cod_piesa:
            producator = lista_producator[code_list.index(cod_piesa)]
    product_details["producator"] = producator
    return product_details


def product_aplicatii(html_file):
    html = open(html_file, "r")
    contents = html.read()
    bs_content = soup(contents, "lxml")
    aplicatii = {
    }
    aplicatii_list = []
    motor = None
    cod_motor = None
    ani_productie = None
    kW = None
    hp = None
    for car_brand_tag in bs_content.find_all("div", class_="tree__branch js-tree-trigger is-open"):
        car_brand = car_brand_tag.text
        next_tag = car_brand_tag.findNext("div")
        for leaf in next_tag.find_all("div", class_="tree__leaf js-tree-trigger is-open"):
            try:
                if leaf.findNext("div").ul["class"] == ['tree__list']:
                    pass
            except:
                name = leaf.text
                tabel = leaf.findNext("div")
                for masina in tabel.find_all("tr", class_="datatable__rowtd datatable__clickable js-clickable-row"):
                    for info in masina.find_all(class_="datatable__item"):
                        try:
                            if info.span.text == "Motor":
                                list_engine = list(info.text)
                                del list_engine[0:30]
                                engine = "".join(list_engine)
                                motor = engine.strip()
                            elif info.span.text == "Codurile motorului":
                                list_engine_code = list(info.text)
                                del list_engine_code[0:30]
                                engine_code = "".join(list_engine_code)
                                cod_motor = engine_code.strip()
                            elif "Anii" in info.span.text:
                                production_years_list = list(info.text)
                                del production_years_list[0:30]
                                production_years = "".join(production_years_list)
                                production_years = production_years.replace("\n", "")
                                production_years = production_years.replace(" ", "")
                                ani_productie = production_years.strip()
                            elif info.span.text == "kW":
                                kw_list = list(info.text)
                                del kw_list[0:30]
                                kw = "".join(kw_list)
                                kW = f"{kw.strip()}kW"
                            elif info.span.text == "CP":
                                hp_list = list(info.text)
                                del hp_list[0:30]
                                horse = "".join(hp_list)
                                hp = f"{horse.strip()}cp"
                        except:
                            pass
                    aplicatii_list.append([motor, cod_motor, hp, kW, ani_productie])
                aplicatii[f"{car_brand} {name}"] = aplicatii_list
                aplicatii_list = []
    return aplicatii


def descriere(aplicatii, product_details, cod_produs):
    oem_equivalent = product_details["oem_equivalent"]
    descriere_tehnica = product_details["descriere_tehnica"]
    tabel_compatibilitate = []
    tabel_echivalente = []
    for key, value in aplicatii.items():
        tabel_compatibilitate.append(f"<div><b>{key}:</b></div>")
        for val in value:
            val_string = " ".join(val)
            tabel_compatibilitate.append(f"<div>{val_string}</div>")
        tabel_compatibilitate.append("<div><br></div>")
    tabel_compatibilitate = " ".join(tabel_compatibilitate)
    for ech in oem_equivalent:
        ech_string = "".join(ech)
        tabel_echivalente.append(f"<div><b>{ech_string}</b></div>")
    tabel_echivalente = " ".join(tabel_echivalente)
    descriere = f"""<h3>{descriere_tehnica}</h3><br>
        <div><br></div>
        <div><br></div>
        <h3><u>Masini compatibile:</u></h3>
        {tabel_compatibilitate}
        <div><br></div>
        <div><br></div>
        <h3>Echivalente coduri original:</h3>
        {tabel_echivalente}
        """
    return descriere

adauga_excel = []
directory = "C:\\Users\\HP\\Desktop\\ALLPARTS\\VOLANT\\intercar_pret_livrare_descr"
for html_file in os.listdir(directory):
    cod_produs = html_file.replace(".html", "")
    print(cod_produs)
    html = f"{directory}\\{html_file}"
    html_cod = open(html, "r")
    content = html_cod.read()
    bs_content = soup(content, "lxml")
    delivery_date = bs_content.find("div", class_="productdelivery__date").text
    if "La cerere" in delivery_date:
        pass
    else:
        product_details = product_data(html, cod_produs)
        aplicatii_produs = product_aplicatii(html)
        descriere_anunt = descriere(aplicatii_produs, product_details, cod_produs)
        tip_produs = product_details["tip_produs"]
        producator = product_details["producator"]
        oem_equivalent = product_details["oem_equivalent"]
        img_src = product_details["img_src"]
        price = product_details["price"]
        for key in aplicatii_produs.keys():
            titlu = f"{tip_produs} {key}  {producator} - {cod_produs}"
            adauga_excel.append([titlu, tip_produs, descriere_anunt, "RON", price, "1", img_src])

workbook = xlsxwriter.Workbook("C:\\Users\\HP\\Desktop\\ALLPARTS\\VOLANT\\VOLANT_ANUNTURI.xlsx")
worksheet = workbook.add_worksheet("Sheet1")
row = 0
for car in adauga_excel:
    print(row)
    worksheet.write(row, 0, car[0])
    worksheet.write(row, 1, car[1])
    worksheet.write(row, 2, car[2])
    worksheet.write(row, 3, car[3])
    worksheet.write(row, 4, car[4])
    worksheet.write(row, 5, car[5])
    worksheet.write(row, 6, car[6])
    row += 1
workbook.close()

